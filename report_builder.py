"""
report_builder.py
------------------
Builds the styled "Visit Views per SA" Excel workbook and converts it into
HTML (via xlsx2html) for the email-ready view and Excel download.
"""

import io
from io import StringIO
from typing import Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from xlsx2html import xlsx2html

from data_logic import (
    CLEAN_CONTACT_PHONE_COL,
    CLEAN_DAYS_TO_INST_COL,
    CLEAN_INSTALMENT_COL,
    COLOR_RED,
    COLOR_YELLOW,
    _days_to_inst_bucket,
)

FONT_FAMILY = "Segoe UI"


def _to_native(value):
    """Convert pandas/numpy scalars to plain Python types openpyxl can write cleanly."""
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d")
    if hasattr(value, "item"):  # numpy / pandas nullable scalar
        try:
            return value.item()
        except Exception:
            return value
    return value


def build_visit_report_workbook(df: pd.DataFrame):
    """
    Build the styled Visit Views workbook.

    Applies:
      - a timestamp row (if report_date is given)
      - bold, filled, bordered header row
      - light red / light yellow row highlighting for accounts due in
        0-2 / 3-5 days
      - peso currency format for Instalment, plain integer format for
        Days to Instalment and Contact Phone

    Returns (workbook, raw_xlsx_bytes, excel_html_content).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visit View"
    ws.views.sheetView[0].showGridLines = True

    thin_side = Side(border_style="thin", color="D9D9D9")
    all_borders = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    timestamp_font = Font(name=FONT_FAMILY, size=10, italic=True, color="555555")
    header_font = Font(name=FONT_FAMILY, size=11, bold=True, color="1B365D")
    data_font = Font(name=FONT_FAMILY, size=11, color="333333")
    header_fill = PatternFill(start_color="EEECE1", end_color="EEECE1", fill_type="solid")
    red_fill = PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type="solid")
    yellow_fill = PatternFill(start_color=COLOR_YELLOW, end_color=COLOR_YELLOW, fill_type="solid")

    current_row = 1
    # if report_date is not None:
    #     label = report_date.strftime("%B %d, %Y") if hasattr(report_date, "strftime") else str(report_date)
    #     ws.cell(row=current_row, column=1, value=f"Report Generated: {label}").font = timestamp_font
    #     current_row += 2
    # else:
    #     current_row += 1

    columns = list(df.columns)

    # Header row
    for col_idx, col_name in enumerate(columns, start=1):
        c = ws.cell(row=current_row, column=col_idx, value=col_name)
        c.font, c.fill, c.border = header_font, header_fill, all_borders
        c.alignment = Alignment(horizontal="left" if col_idx == 1 else "center", vertical="center")
    header_row_idx = current_row
    current_row += 1

    instalment_col_num = columns.index(CLEAN_INSTALMENT_COL) + 1 if CLEAN_INSTALMENT_COL in columns else None
    days_col_num = columns.index(CLEAN_DAYS_TO_INST_COL) + 1 if CLEAN_DAYS_TO_INST_COL in columns else None
    phone_col_num = columns.index(CLEAN_CONTACT_PHONE_COL) + 1 if CLEAN_CONTACT_PHONE_COL in columns else None

    # Data rows
    for _, row in df.iterrows():
        bucket = (
            _days_to_inst_bucket(row[CLEAN_DAYS_TO_INST_COL])
            if CLEAN_DAYS_TO_INST_COL in df.columns
            else None
        )
        fill = red_fill if bucket == "red" else yellow_fill if bucket == "yellow" else None

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=_to_native(row[col_name]))
            cell.font = data_font
            cell.border = all_borders
            if fill is not None:
                cell.fill = fill

            if col_idx == instalment_col_num:
                cell.number_format = '"\u20b1"#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in (days_col_num, phone_col_num):
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

    # Auto column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None and cell.row >= header_row_idx:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

    xlsx_buffer = io.BytesIO()
    wb.save(xlsx_buffer)
    xlsx_buffer.seek(0)
    raw_xlsx_bytes = xlsx_buffer.getvalue()

    excel_html_buffer = StringIO()
    xlsx2html(xlsx_buffer, excel_html_buffer)
    excel_html_content = excel_html_buffer.getvalue()

    excel_html_content = excel_html_content.replace('&quot;₱&quot;', '₱')
    excel_html_content = excel_html_content.replace('"₱"', '₱')

    return wb, raw_xlsx_bytes, excel_html_content


def table_to_email(table_html, agent='Agent', report_signature=None):
    if not report_signature:
        report_signature = 'MSME Team'
    """Assembles the final email HTML."""
    email_template = f"""
    <!DOCTYPE html>
    <html lang="tl">
    <head>
    <meta charset="UTF-8">
    <title>Paalala: Customer Visits</title>
    </head>
    <body>
    <div class="container">
    <b>Magandang araw, {agent}! 🙂</b>

    <p>Salamat sa iyong suporta at pag-aasikaso sa ating mga customer. Malaking tulong ang ginagawa mo para mas makapagserbisyo tayo sa kanila.</p>

    <p>Para mas mapatatag pa natin ang relasyon sa ating mga kliyente, pinaaalala namin ang <b><u> personal mong pagbisita sa mga customer</u></b> na nasa attachment dahil may paparating silang due dates.</p>

    <div class="highlight">
        Para sa buwan na ito, <b><u> mandatory ang face to face visits at ito ang dapat unahin</u></b>. 
        Dapat mong dalawin ang kanilang mga tindahan, kumustahin sila, at kausapin tungkol sa kanilang nalalapit na bayarin.
    </div>

    <span style="color: red;"><p><b>
        Ikaw ay <strong>sales partner</strong> ng iyong mga customer. 
        Ang tungkulin mo ay maayos at magalang na magpaalala na malapit na ang due nila. 
        <u>Hindi mo kailangan</u> o dapat singilin sila kapag nahuli sila sa bayad. 
        May hiwalay na team ang Home Credit para sa ganitong mga sitwasyon.
    </b></p></span>

    <p>Kung sakaling hindi agad posible ang face to face visit, maaari mo silang tawagan bilang pangalawang opsyon.</p>

    <p><b>Sa pagbisita o pagtawag, pakisuri ang mga sumusunod:</b></p>
    <ul>
        <li>Kumusta ang customer</li>
        <li>Handa ba silang magbayad sa tamang oras</li>
        <li>Kailangan ba nila ng tulong sa pag-navigate ng proseso ng pagbabayad</li>
        <li>Kailangan ba nila ng tulong para makontak ang customer support</li>
    </ul>

    <p>Naka-attach ang kanilang mga pangalan at contract details para sa iyong gabay.</p>

    <div style="overflow-x:auto;">
    {table_html}
    </div>

    <p><b><u>
        Pagkatapos mong makapag-check in, pakifill out ang short survey 
        <a href="https://forms.office.com/pages/responsepage.aspx?id=28KJynnWnUiJ-ySGzZre1O5YneIZopZOjWlglhamDPlUMTJQMlZHRUhGSDhZMVFWWEhCNkxVUVNRViQlQCN0PWcu">(MS Form link)</a> 
        para mas ma-monitor at masuportahan namin ang iyong mga follow up.
    </u></b></p>

    <p>Sana makatulong ito para mas maging maayos at mas maging malapit pa ang pakikipag-ugnayan mo sa ating mga customer.</p>

    <p class="footer">Maraming salamat, at ingat ka palagi! 😊<br>

    <strong>{report_signature}</strong></p>
    </div>
    </body>
    </html>

    """
    return email_template